from collections import OrderedDict
import csv
import definitions
from openpyxl.worksheet.read_only import ReadOnlyWorksheet
import os
import re

YEAR_SEP = ' --- '


def remove_extra_whitespace(v):
    return re.sub(r'\s+', ' ', v)


def prepare_str(v):
    if v is None:
        v = ''

    v = str(v).strip()
    return remove_extra_whitespace(v)


def get_table_num(sheet_name, regex=r'^Table (\d+)'):
    m = re.search(regex, sheet_name)
    if m is None:
        raise ValueError('"{0}" has no number!'.format(sheet_name))
    num = m.group(1)
    return num


def get_table_num_from_sheet(sheet):
    assert isinstance(sheet, ReadOnlyWorksheet), 'Found type {0}'.format(type(sheet))
    return get_table_num(sheet.title)


def get_header_rows(sheet):
    num = get_table_num_from_sheet(sheet)
    ci = definitions.CONVERT_INFO[num]

    header_rows = ci.header_rows
    year_row = ci.year_row
    header = ci.header.strip()

    if not header_rows:  # use the header if header_rows row is -1 is nothing
        assert year_row, 'Must have a year row when header_rows not specified'
        assert header, 'Header needs to be provided if header_rows is not!'

    if year_row and year_row not in header_rows:
        header_rows = [year_row] + header_rows

    return year_row, header, header_rows


def split_header(header, header_sep=', '):
    return header.split(header_sep)


def has_year(header):
    return YEAR_SEP in header


def remove_year(header):
    if has_year(header):
        return header.split(YEAR_SEP)[1]
    else:
        return header


def add_year(header, year):
    return '{0}{2}{1}'.format(year, header, YEAR_SEP)


def get_input_headers(sheet, header_sep=', ', year_sep=' - '):
    assert isinstance(sheet, ReadOnlyWorksheet), 'Found type {0}'.format(type(sheet))

    year_row, header, header_rows = get_header_rows(sheet)

    years = []
    headers = []
    year = 0
    columns = OrderedDict()
    curr_data = [''] * len(header_rows)  # assume blank unless filled
    for c in range(1, sheet.max_column + 1):
        updated = False

        if c != 1 and year_row:
            year_val = sheet.cell(row=year_row, column=c).value
            if year_val:
                year = int(year_val)
                if year not in years:
                    years.append(year)

        for i, r in enumerate(header_rows):

            v = str(sheet.cell(row=r, column=c).value or '').strip()
            if c != 1 and r == year_row:

                if len(header_rows) == 1 and not v:
                    updated = False
                    break

                curr_data[i] = header.strip()
                updated = True

            elif v:

                curr_data[i] = prepare_str(v)
                updated = True

        if not updated:
            break

        header_parts = [d for d in curr_data if d]

        if header_parts:

            i = curr_data.index(header_parts[-1])
            curr_data[i:] = [''] * (len(header_rows) - i)

            h = header_sep.join(header_parts)
            if h not in headers:
                headers.append(h)

            # key = add_year(h, year) if year_row else h
            if h not in columns:
                columns[h] = OrderedDict()
            columns[h][year] = c - 1  # maps to zero-based column number

    for h, v in columns.items():
        if 0 in v:
            for y in years:
                if y not in v:
                    v[y] = v[0]

    return years, headers, columns


def get_year_info(sheet):
    num = get_table_num_from_sheet(sheet)
    return definitions.YEAR_INFO[num]


def get_year_by_cell(cell, regex, default=None):
    m = re.search(regex, str(cell.value or ''))
    if m is None:
        if default is None:
            raise ValueError('No matching value for year in cell: {0}'.format(cell))
        return int(default)
    return int(m.group(1))


def get_year(sheet, rows, default=None):
    year_info = get_year_info(sheet)

    if len(year_info) == 0:
        return 0

    if 'cell' in year_info:
        cell = sheet[year_info['cell']]
    elif 'column' in year_info:
        cell = rows[year_info['column']]
    else:
        raise ValueError('Must have either "cell" or "column" in YEAR_INFO!')

    regex = year_info['regex']

    return int(get_year_by_cell(cell, regex, default=default))


def is_number(v):
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def mkdirp(file_dir):
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    return file_dir


def mkfdirp(file_path):
    mkdirp(os.path.dirname(file_path))
    return file_path


def write_csv(out_file, headers, data):

    # make directory
    mkfdirp(out_file)

    # write to temporary file
    with open(out_file + '.tmp', 'w', newline='\n') as f:
        cw = csv.DictWriter(f, fieldnames=headers)
        cw.writeheader()
        for row in data:
            cw.writerow(row)

    # move temporary file to actual location
    if os.path.exists(out_file):
        os.remove(out_file)
    os.rename(out_file + '.tmp', out_file)


def merge_dicts(d, *others):
    o = d.copy()
    for oi in others:
        o.update(oi)
    return o
