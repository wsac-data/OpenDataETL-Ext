"""
Prepared for Chadd in response to Upwork posting

Developed by Andre Dozier 2018

Converts
"""

import argparse
from collections import OrderedDict
import convert
import definitions
import itertools
from openpyxl import load_workbook
from openpyxl.worksheet.read_only import ReadOnlyWorksheet
import re
from time import time
import utils
import warnings


def retrieve_data(workbook, pattern, verbose=False):

    wb = load_workbook(workbook, read_only=True)
    sheets = [n for n in wb.sheetnames if re.search(pattern, n, flags=re.I)]

    out_headers = []
    out_data = OrderedDict()

    for sheet_i, sheet_name in enumerate(sheets):

        if verbose:
            print('  Sheet "{0}"...'.format(sheet_name))

        sheet = wb[sheet_name]
        assert isinstance(sheet, ReadOnlyWorksheet), 'Found type {0}'.format(type(sheet))

        year_row, _, header_rows = utils.get_header_rows(sheet)
        years, headers, columns = utils.get_input_headers(sheet)
        default_cvt = convert.get_default_converter_by_sheet(sheet)

        # headers
        curr_headers = convert.get_headers(sheet)

        if sheet_i == 0:
            out_headers = curr_headers
        else:
            if curr_headers != out_headers:
                warnings.warn('Differences:\n  {0}'.format('\n  '.join(set(out_headers) - set(curr_headers))))

        def get_col(h, yr):
            d = columns[h]
            if isinstance(yr, list):
                if yr:
                    yr = yr[0]
                else:
                    yr = 0
            else:
                yr = int(yr)
            return d[yr] if yr in d else next(v for v in d.values())

        char_col = get_col(headers[0], years)
        check_col = get_col(headers[1], years)

        cvt = default_cvt

        year = 0
        characteristic = ''
        sub_char = ''

        if year_row:
            gen_years = years
        else:
            gen_years = [None]

        for y in gen_years:

            for row in itertools.islice(sheet.rows, max(header_rows), None):

                if not year_row:
                    year = utils.get_year(sheet, row, default=year)
                    y = year

                check_val = utils.prepare_str(row[check_col].value)
                char_val = utils.prepare_str(row[char_col].value)

                if char_val:
                    sub_char = char_val

                if utils.is_number(check_val) or check_val in {'*', '†', 'n.a.'}:

                    key = '__'.join([str(y), characteristic, utils.prepare_str(row[char_col].value)])

                    if key not in out_data:
                        out_data[key] = {'Sub-Characteristic': sub_char}

                        if characteristic and 'Characteristic' in out_headers:
                            out_data[key]['Characteristic'] = characteristic

                        if y:
                            out_data[key]['Year'] = str(y)

                    # add data to row
                    out_data[key].update({new_h: c(row[get_col(h, y)].value)
                                          for h, new_h, c in cvt.iter_headers(headers[1:])})

                else:

                    characteristic = char_val
                    if not characteristic and check_val:
                        cvt = convert.DataTypeConverter(check_val, default=default_cvt)

    return out_headers, out_data


def transform(workbook, pattern, output, beg_cols=(), verbose=False):
    """
    Transform data within a messy Excel workbook file

    :param workbook: Workbook file path
    :param pattern: Pattern of sheet names to extract
    :param output: File path to CSV output file
    :param beg_cols: Beginning columns
    :param verbose: Verbosity
    :return: Returns transformed workbook
    """

    print('Working on {0}...'.format(output))
    start = time()
    headers, data = retrieve_data(workbook, pattern, verbose=verbose)
    headers = [c for c, v in beg_cols] + headers
    out_data = [utils.merge_dicts(v, beg_cols) for v in data.values()]
    utils.write_csv(output, headers, out_data)
    print('  Finished in {0:.1f} seconds'.format(time() - start))


def transform_all(workbook, pattern, output, beg_cols=(), verbose=False):

    def t(s):
        return s.title()

    if '{source}' in workbook or '{dollar}' in workbook:
        for source, dollar in definitions.TYPES:
            wb_file = workbook.format(source=source, dollar=dollar)
            out_file = output.format(source=source, dollar=dollar)
            beg_col_fmt = [(
                c.format(source=t(source), dollar=t(dollar)),
                v.format(source=t(source), dollar=t(dollar)),
            ) for c, v in beg_cols]
            transform(wb_file, pattern, out_file, beg_cols=beg_col_fmt, verbose=verbose)

    else:
        transform(workbook, pattern, output, beg_cols=beg_cols, verbose=verbose)


def main():
    parser = argparse.ArgumentParser(description='Transform data')

    parser.add_argument('workbook', type=str, help='Path to the workbook')
    parser.add_argument('pattern', type=str, help='REGEX pattern of the sheet names to extract')
    parser.add_argument('output', type=str, help='Location of CSV output file')
    parser.add_argument('--beg_col', action='append', type=str, nargs=2, default=[],
                        help='Added columns on the left side')
    parser.add_argument('--verbose', action='store_true', default=False, help='Verbosity')

    args = parser.parse_args()

    transform_all(args.workbook, args.pattern, args.output, beg_cols=args.beg_col, verbose=args.verbose)


if __name__ == '__main__':
    main()
